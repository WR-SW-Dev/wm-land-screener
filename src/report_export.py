"""Branded PDF/Excel exports of the qualifying-parcels table.

`build_pdf_report` uses a curated, print-friendly column set (the report's
original spec — address/owner/acres/zone/pathway/units/flood/wetland/score).
`build_excel_report` exports the full table as displayed on screen — no
print-width constraint, so no need to curate.
"""
import io
from datetime import datetime

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape, portrait
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (Image as RLImage, Paragraph, SimpleDocTemplate,
                                Spacer, Table, TableStyle)
from PIL import Image as PILImage

from scoring import SCORE_COMPONENTS

WR_TEAL = "#779FA1"
WR_DARK = "#2c3e3f"
_STRIPE = "#eef4f4"
_BAR_BG = "#e5e7eb"

# Duplicated from app.py (not imported, to avoid a report_export -> app ->
# report_export circular import) — keep in sync if those change.
_SCORE_HIGH, _SCORE_MED = 70, 50
_COLOR_HIGH, _COLOR_MED, _COLOR_LOW = "#22c55e", "#f59e0b", "#ef4444"
_STATUS_COLORS = {
    "Not contacted": "#9ca3af", "Pursuing": "#22c55e",
    "Backburner": "#f59e0b", "No": "#ef4444",
}
_PATHWAY_COLORS = {
    "By right": "#22c55e", "PRD special use": "#3b82f6", "PUD special use": "#3b82f6",
    "Master plan upzone": "#8b5cf6", "PD rezoning": "#f59e0b", "Not viable": "#9ca3af",
}


def _score_color(score):
    if score >= _SCORE_HIGH:
        return _COLOR_HIGH
    if score >= _SCORE_MED:
        return _COLOR_MED
    return _COLOR_LOW

PDF_COLUMNS = [
    ("address", "Address"),
    ("parcel_id", "Parcel #"),
    ("calc_acres", "Acres"),
    ("zone_code", "Zone"),
    ("dev_pathway", "Dev Pathway"),
    ("units_range", "Units (Cons.–Opt.)"),
    ("future_lu_label", "Future Zoning"),
    ("future_units_range", "Future Units (Cons.–Opt.)"),
    ("flood_pct", "Flood %"),
    ("wetland_pct", "Wetland %"),
    ("score", "Score"),
]
_UNITS_COL_IDX = [i for i, (col, _) in enumerate(PDF_COLUMNS) if col == "units_range"][0]
_FUTURE_UNITS_COL_IDX = [i for i, (col, _) in enumerate(PDF_COLUMNS)
                         if col == "future_units_range"][0]

# Same conservative/optimistic multipliers as scoring.py's units_conservative/
# units_optimistic — applied here to future_max_units instead of the current
# zoning's max_units_per_acre, so "future units" is computed the same way as
# "current units", just under the master-plan density instead of today's.
_CONSERVATIVE_MULT = 0.70
_OPTIMISTIC_MULT = 1.00

# Max height for the per-parcel card's aerial thumbnail. Height used to be
# derived purely from width * aspect ratio, so a narrow/tall parcel (common
# for vacant rural land — flag lots, long rear parcels) could produce an
# image several inches taller than a typical wide parcel's, blowing the
# one-page budget and bumping the whole card to page 2. Capping height and
# deriving width from it when that's the binding dimension guarantees the
# image can't do that, at the cost of rendering narrower (not full-column-
# width) for those same odd-shaped parcels.
_MAX_CARD_IMG_HEIGHT = 3.25 * inch


def _pdf_cell(col, v):
    if pd.isna(v):
        return ""
    if col == "calc_acres":
        return f"{float(v):.2f}"
    return str(v)


def _pdf_rows(df):
    d = df.copy()
    if "units_conservative" in d.columns and "units_optimistic" in d.columns:
        d["units_range"] = (d["units_conservative"].astype(str) + "–"
                             + d["units_optimistic"].astype(str))
    if "future_max_units" in d.columns and "net_dev_acres" in d.columns:
        net = pd.to_numeric(d["net_dev_acres"], errors="coerce").fillna(0)
        future_density = pd.to_numeric(d["future_max_units"], errors="coerce").fillna(0)
        cons = (net * future_density * _CONSERVATIVE_MULT).round(0).astype(int)
        opt = (net * future_density * _OPTIMISTIC_MULT).round(0).astype(int)
        d["future_units_range"] = cons.astype(str) + "–" + opt.astype(str)
    rows = [[label for _, label in PDF_COLUMNS]]
    for _, r in d.iterrows():
        rows.append([_pdf_cell(col, r.get(col)) for col, _ in PDF_COLUMNS])
    return rows


def _pdf_header(city_label, count, page_width):
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("WRTitle", parent=styles["Title"],
                                 textColor=colors.white, fontSize=16, leading=19)
    sub_style = ParagraphStyle("WRSub", parent=styles["Normal"],
                               textColor=colors.white, fontSize=9)
    text_cell = [
        Paragraph(f"{city_label} — Qualified Parcels Report", title_style),
        Paragraph(f"{count} parcels · Generated {datetime.now():%B %d, %Y}", sub_style),
    ]
    header = Table([[text_cell]], colWidths=[page_width])
    header.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(WR_DARK)),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (0, 0), 14),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
    ]))
    return header


def build_pdf_report(df, city_label: str) -> bytes:
    buf = io.BytesIO()
    page_size = landscape(letter)
    doc = SimpleDocTemplate(buf, pagesize=page_size,
                            topMargin=0.4 * inch, bottomMargin=0.5 * inch,
                            leftMargin=0.5 * inch, rightMargin=0.5 * inch)
    content_width = page_size[0] - doc.leftMargin - doc.rightMargin

    elements = [_pdf_header(city_label, len(df), content_width), Spacer(1, 14)]

    rows = _pdf_rows(df)
    table = Table(rows, repeatRows=1)
    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(WR_TEAL)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cccccc")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("ALIGN", (_UNITS_COL_IDX, 0), (_UNITS_COL_IDX, -1), "CENTER"),
        ("ALIGN", (_FUTURE_UNITS_COL_IDX, 0), (_FUTURE_UNITS_COL_IDX, -1), "CENTER"),
    ]
    for i in range(2, len(rows), 2):
        style_cmds.append(("BACKGROUND", (0, i), (-1, i), colors.HexColor(_STRIPE)))
    table.setStyle(TableStyle(style_cmds))
    elements.append(table)

    def _footer(canvas, doc_):
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(colors.grey)
        canvas.drawString(0.5 * inch, 0.3 * inch,
                          "WR-Dev internal screening tool — estimates only; "
                          "verify zoning and unit counts before use.")
        canvas.drawRightString(page_size[0] - 0.5 * inch, 0.3 * inch, f"Page {doc_.page}")
        canvas.restoreState()

    doc.build(elements, onFirstPage=_footer, onLaterPages=_footer)
    return buf.getvalue()


def build_excel_report(df, city_label: str) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Qualified Parcels"

    n_cols = max(len(df.columns), 1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    title_cell = ws.cell(row=1, column=1, value=f"{city_label} — Qualified Parcels Report")
    title_cell.font = Font(size=14, bold=True, color="FFFFFF")
    for col in range(1, n_cols + 1):
        ws.cell(row=1, column=col).fill = PatternFill("solid", fgColor=WR_DARK.lstrip("#"))
    ws.row_dimensions[1].height = 26

    sub_cell = ws.cell(row=2, column=1,
                       value=f"{len(df)} parcels · Generated {datetime.now():%B %d, %Y}")
    sub_cell.font = Font(italic=True, size=9, color="666666")

    header_row = 4
    for j, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=header_row, column=j, value=str(col))
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=WR_TEAL.lstrip("#"))
        cell.alignment = Alignment(horizontal="center")

    for i, (_, r) in enumerate(df.iterrows(), start=header_row + 1):
        for j, col in enumerate(df.columns, start=1):
            v = r[col]
            ws.cell(row=i, column=j, value=(None if pd.isna(v) else v))
        if (i - header_row) % 2 == 0:
            for j in range(1, n_cols + 1):
                ws.cell(row=i, column=j).fill = PatternFill("solid", fgColor=_STRIPE.lstrip("#"))

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate
    for j, col in enumerate(df.columns, start=1):
        ws.column_dimensions[get_column_letter(j)].width = max(10, min(28, len(str(col)) + 6))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Per-parcel listing card — one page, modeled on a CARWM-style comp sheet ────
def _v(row, col, default=None):
    val = row.get(col) if hasattr(row, "get") else None
    return default if val is None or pd.isna(val) else val


def _badge(text, bg_hex, width=None, font_size=9):
    t = Table([[text]], colWidths=[width] if width else None)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(bg_hex)),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.white),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), font_size),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
    ]))
    return t


def _bar(pct, color_hex, width=1.7 * inch, height=8):
    pct = max(0.0, min(1.0, pct))
    filled = max(1, round(width * pct))
    t = Table([["", ""]], colWidths=[filled, width - filled], rowHeights=[height])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, 0), colors.HexColor(color_hex)),
        ("BACKGROUND", (1, 0), (1, 0), colors.HexColor(_BAR_BG)),
        ("TOPPADDING", (0, 0), (-1, -1), 0), ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ("LEFTPADDING", (0, 0), (-1, -1), 0), ("RIGHTPADDING", (0, 0), (-1, -1), 0),
    ]))
    return t


def _fact_box(title, rows, styles, box_width):
    body = [[Paragraph(f"<b>{title}</b>", styles["FactTitle"])]]
    for label, value in rows:
        body.append([Paragraph(f"<font color='#888888'>{label}</font>  {value}", styles["Fact"])])
    t = Table(body, colWidths=[box_width])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(WR_TEAL)),
        ("TOPPADDING", (0, 0), (-1, -1), 5), ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 8), ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#cccccc")),
        ("LINEBELOW", (0, 0), (-1, 0), 0.5, colors.HexColor("#cccccc")),
    ]))
    return t


def _overview_text(row):
    acres = _v(row, "calc_acres", 0) or 0
    net = _v(row, "net_dev_acres", acres) or acres
    zone_c = _v(row, "zone_code", "") or ""
    zone_l = _v(row, "zone_label", "") or ""
    pathway = _v(row, "dev_pathway", "") or ""
    u_con = int(_v(row, "units_conservative", 0) or 0)
    u_opt = int(_v(row, "units_optimistic", 0) or 0)
    flood = float(_v(row, "flood_pct", 0) or 0) * 100
    wetland = float(_v(row, "wetland_pct", 0) or 0) * 100
    flu_label = _v(row, "future_lu_label", "") or ""
    flu_max = _v(row, "future_max_units", None)
    rezone_up = bool(_v(row, "rezoning_upside", False))

    parts = [
        f"This {acres:.2f}-acre parcel is zoned {zone_c}"
        f"{f' ({zone_l})' if zone_l else ''}"
        f"{f', with a development pathway of {pathway.lower()}' if pathway else ''}."
    ]
    if net and u_opt:
        parts.append(
            f"Under current zoning, the {net:.2f} net developable acres could "
            f"support an estimated {u_con}–{u_opt} units."
        )
    if flood > 0:
        parts.append(f"Approximately {flood:.1f}% of the site falls within a mapped floodplain.")
    if wetland > 0:
        parts.append(f"Approximately {wetland:.1f}% of the site is mapped wetland.")
    if flu_label:
        upside = " — an increase over current zoning" if rezone_up else ""
        parts.append(
            f"The area's master-plan future land use is designated {flu_label}"
            f"{f', supporting up to {int(flu_max)} units/acre' if flu_max else ''}{upside}."
        )
    return " ".join(parts)


def build_parcel_card(row, city_label: str, thumbnail_png: bytes = None,
                      status: str = "Not contacted", notes: str = "") -> bytes:
    """One-page listing card for a single parcel — CARWM-style: header with
    address/status, aerial thumbnail + overview, a key-facts sidebar, and a
    score breakdown. `row` is a dict-like of that parcel's raw pipeline
    columns (same fields already used in the map popup — no new data)."""
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle("CardAddr", fontSize=16, fontName="Helvetica-Bold",
                              textColor=colors.white, leading=19))
    styles.add(ParagraphStyle("CardOwner", fontSize=10, textColor=colors.HexColor("#cbd5c8")))
    styles.add(ParagraphStyle("CardScore", fontSize=22, leading=26, fontName="Helvetica-Bold",
                              textColor=colors.white, alignment=2))
    styles.add(ParagraphStyle("CardScoreSub", fontSize=8, leading=11,
                              textColor=colors.HexColor("#cbd5c8"), alignment=2))
    styles.add(ParagraphStyle("SectionHead", fontSize=11, fontName="Helvetica-Bold",
                              textColor=colors.HexColor(WR_DARK), spaceAfter=4))
    styles.add(ParagraphStyle("Body", fontSize=9.5, leading=13))
    styles.add(ParagraphStyle("Notes", fontSize=9, leading=12, fontName="Helvetica-Oblique",
                              textColor=colors.HexColor("#555555")))
    styles.add(ParagraphStyle("FactTitle", fontSize=10, fontName="Helvetica-Bold",
                              textColor=colors.white))
    styles.add(ParagraphStyle("Fact", fontSize=9, leading=13))

    buf = io.BytesIO()
    page_size = portrait(letter)
    doc = SimpleDocTemplate(buf, pagesize=page_size,
                            topMargin=0.4 * inch, bottomMargin=0.5 * inch,
                            leftMargin=0.5 * inch, rightMargin=0.5 * inch)
    content_width = page_size[0] - doc.leftMargin - doc.rightMargin
    left_width = content_width * 0.6
    right_width = content_width - left_width - 0.2 * inch

    address = _v(row, "address", "(no address)") or "(no address)"
    owner = _v(row, "owner", "") or ""
    score = float(_v(row, "score", 0) or 0)
    status_badge = _badge(status, _STATUS_COLORS.get(status, "#9ca3af"))
    header_left = [Paragraph(address, styles["CardAddr"])]
    if owner:
        header_left.append(Paragraph(owner, styles["CardOwner"]))
    header_left.append(Spacer(1, 4))
    header_left.append(status_badge)
    header_right = [Paragraph(f"{score:.0f}", styles["CardScore"]),
                    Paragraph("/ 100 score", styles["CardScoreSub"])]
    header = Table([[header_left, header_right]],
                   colWidths=[content_width * 0.7, content_width * 0.3])
    header.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(WR_DARK)),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (0, 0), 0.5 * inch),
        ("RIGHTPADDING", (1, 0), (1, 0), 0.5 * inch),
        ("TOPPADDING", (0, 0), (-1, -1), 16),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 16),
    ]))

    # ── Left column: aerial thumbnail + overview + notes ──────────────────────
    left = []
    if thumbnail_png:
        img_buf = io.BytesIO(thumbnail_png)
        pil_img = PILImage.open(img_buf)
        aspect = pil_img.height / pil_img.width   # height/width
        img_buf.seek(0)
        img_w, img_h = left_width, left_width * aspect
        if img_h > _MAX_CARD_IMG_HEIGHT:           # narrow/tall parcel — cap height, derive width
            img_h = _MAX_CARD_IMG_HEIGHT
            img_w = img_h / aspect
        img_cell = Table([[RLImage(img_buf, width=img_w, height=img_h)]],
                         colWidths=[left_width])
        img_cell.setStyle(TableStyle([
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("LEFTPADDING", (0, 0), (-1, -1), 0), ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 0), ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ]))
        left.append(img_cell)
        left.append(Spacer(1, 10))
    left.append(Paragraph("Overview", styles["SectionHead"]))
    left.append(Paragraph(_overview_text(row), styles["Body"]))
    if notes:
        left.append(Spacer(1, 8))
        left.append(Paragraph(f"Notes: {notes}", styles["Notes"]))

    # ── Right column: key facts / future zoning / score breakdown ────────────
    pathway = _v(row, "dev_pathway", "") or ""
    u_con = int(_v(row, "units_conservative", 0) or 0)
    u_opt = int(_v(row, "units_optimistic", 0) or 0)
    facts_rows = [
        ("Parcel #", _v(row, "parcel_id", "—")),
        ("Acres", f"{float(_v(row, 'calc_acres', 0) or 0):.2f} gross / "
                  f"{float(_v(row, 'net_dev_acres', 0) or 0):.2f} net"),
        ("Zone", f"{_v(row, 'zone_code', '—')} — {_v(row, 'zone_label', '')}"),
        ("Dev pathway", pathway or "—"),
        ("Units (cons.–opt.)", f"{u_con}–{u_opt}"),
        ("Flood %", f"{float(_v(row, 'flood_pct', 0) or 0) * 100:.1f}%"),
        ("Wetland %", f"{float(_v(row, 'wetland_pct', 0) or 0) * 100:.1f}%"),
        ("MF / ADU", f"{_v(row, 'mf_permitted', '—')} / {_v(row, 'adu_permitted', '—')}"),
    ]
    right = [_fact_box("Key facts", facts_rows, styles, right_width)]

    flu_label = _v(row, "future_lu_label", "")
    if flu_label:
        flu_max = int(_v(row, "future_max_units", 0) or 0)
        net = float(_v(row, "net_dev_acres", 0) or 0)
        f_con = round(net * flu_max * _CONSERVATIVE_MULT)
        f_opt = round(net * flu_max * _OPTIMISTIC_MULT)
        right.append(Spacer(1, 10))
        right.append(_fact_box("Future zoning", [
            ("Designation", flu_label),
            ("Future units (cons.–opt.)", f"{f_con}–{f_opt}"),
        ], styles, right_width))

    score_rows = [[Paragraph("<b>Score breakdown</b>", styles["FactTitle"])]]
    pts_rezoning_earned = 0.0
    for comp in SCORE_COMPONENTS:
        pts = min(float(_v(row, comp["key"], 0) or 0), comp["max"])
        if comp.get("bonus"):
            pts_rezoning_earned = pts
            continue
        pct = pts / comp["max"] if comp["max"] else 0
        bar_color = _COLOR_HIGH if pct >= 0.8 else (_COLOR_MED if pct >= 0.4 else _COLOR_LOW)
        score_rows.append([Paragraph(f"{comp['label']} — {pts:.0f}/{comp['max']}",
                                     styles["Fact"])])
        score_rows.append([_bar(pct, bar_color, width=right_width - 16)])
    if pts_rezoning_earned > 0:
        score_rows.append([Paragraph(f"Rezoning bonus — +{pts_rezoning_earned:.0f}",
                                     styles["Fact"])])
        score_rows.append([_bar(pts_rezoning_earned / 10, WR_TEAL, width=right_width - 16)])
    score_box = Table(score_rows, colWidths=[right_width])
    score_box.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(WR_TEAL)),
        ("TOPPADDING", (0, 0), (-1, -1), 3), ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 8), ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#cccccc")),
        ("LINEBELOW", (0, 0), (-1, 0), 0.5, colors.HexColor("#cccccc")),
    ]))
    right.append(Spacer(1, 10))
    right.append(score_box)

    body_table = Table([[left, right]], colWidths=[left_width, right_width])
    body_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (0, 0), 0), ("RIGHTPADDING", (0, 0), (0, 0), 0.2 * inch),
        ("LEFTPADDING", (1, 0), (1, 0), 0), ("RIGHTPADDING", (1, 0), (1, 0), 0),
    ]))

    elements = [header, Spacer(1, 16), body_table]

    def _footer(canvas, doc_):
        canvas.saveState()
        canvas.setStrokeColor(colors.HexColor(WR_TEAL))
        canvas.setLineWidth(2)
        canvas.line(0.5 * inch, 0.42 * inch, page_size[0] - 0.5 * inch, 0.42 * inch)
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(colors.grey)
        canvas.drawString(0.5 * inch, 0.28 * inch,
                          f"{city_label} — WR-Dev internal screening tool. Estimates "
                          f"only; verify zoning and unit counts before use.")
        canvas.drawRightString(page_size[0] - 0.5 * inch, 0.28 * inch,
                               f"Generated {datetime.now():%B %d, %Y}")
        canvas.restoreState()

    doc.build(elements, onFirstPage=_footer, onLaterPages=_footer)
    return buf.getvalue()
